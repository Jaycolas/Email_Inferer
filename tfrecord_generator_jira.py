#!/usr/bin/python
# -*- coding: UTF-8 -*-

import glob
import os
import tensorflow as tf
import io
from util import Vocab, tokenize_helper
import pickle
import random
import openpyxl

__author__ = 'Jaycolas'

FILEPATH='./dataset/jira'
XLSX_PATH='./XLSX'
JIRA_WORKBOOK = 'Qualcomm_Custom_JIRA.xlsx'
input_fname_pattern = '*.xlsx'

TRAIN_TFRECORD_FILE = os.path.join(FILEPATH, 'train.tfrecords')
DEV_TFRECORD_FILE = os.path.join(FILEPATH, 'dev.tfrecords')
VAL_TFRECORD_FILE = os.path.join(FILEPATH, 'val.tfrecords')
DEV_SAMPLE_PER = 0.2
VAL_SAMPLE_PER = 0.1
SUMMARY_COL='A'
ASSIGNEE_COL='B'
MPSS_PL_COL='D'
DESCRIPTION_COL='E'
#doc_file_list = glob.glob(os.path.join(FILEPATH, input_fname_pattern))
jira_xlxs_list = glob.glob(os.path.join(XLSX_PATH, input_fname_pattern))
LOWER_DIC_FILTER_THRESHOLD = 13


def save_obj(filepath, obj, name):
    with open(filepath + '/vocab/'+ name + '.pkl', 'wb') as f:
        pickle.dump(obj, f, pickle.HIGHEST_PROTOCOL)


def load_obj(filepath, name):
    with open(filepath + '/vocab/' + name + '.pkl', 'rb') as f:
        return pickle.load(f)


def buildVocab(work_sheet_list):
    print "Building vocabulary for JIRA Fields"

    # general vocabulary is for all jira fields input
    general_vocab = Vocab()
    # assignee_vocab only includes the assignee, which can be used for model which only use assignee as input
    assignee_vocab = Vocab()
    mpss_pl_vocab = Vocab()

    for jira_workbook in work_sheet_list:
        print "Openning ", jira_workbook
        jira_wb = openpyxl.load_workbook(jira_workbook)
        sheet_name = jira_wb.sheetnames[0]
        #print "Sheet name is ", sheet_name
        work_sheet = jira_wb[sheet_name]

        summary_column = work_sheet[SUMMARY_COL]
        assignee_column = work_sheet[ASSIGNEE_COL]
        #status_column = work_sheet['F']
        description_column = work_sheet[DESCRIPTION_COL]
        mpss_pl_column = work_sheet[MPSS_PL_COL]

        for cell in description_column:
            preprocessed_description = tokenize_helper(cell.value)
            if preprocessed_description:
                general_vocab.construct(preprocessed_description.split())

        for cell in summary_column:
            preprocessed_description = tokenize_helper(cell.value)
            if preprocessed_description:
                general_vocab.construct(preprocessed_description.split())

        #for cell in status_column:
            #preprocessed_description = tokenize_helper(cell.value)
            #if preprocessed_description:
                #general_vocab.construct(preprocessed_description.split())

        for cell in assignee_column:
            preprocessed_description = tokenize_helper(cell.value)
            if preprocessed_description:
                general_vocab.construct(preprocessed_description.split())
                #Note that for assignee, we don't need to split the name of each
                assignee_vocab.construct([cell.value])

        for cell in mpss_pl_column:
            preprocessed_description = tokenize_helper(cell.value)
            if preprocessed_description:
                general_vocab.construct(preprocessed_description.split())
                mpss_pl_vocab.construct(preprocessed_description.split())

    save_obj(FILEPATH, general_vocab, 'general_vocab')
    save_obj(FILEPATH, mpss_pl_vocab, 'mpss_pl_vocab')
    save_obj(FILEPATH, assignee_vocab, 'assignee_vocab')


def split_train_dev_val(work_sheet_list, dev_per, val_per):
    #Firstly need to check the validity of each input percentage.
    assert dev_per>0 and dev_per<1
    assert val_per>0 and val_per<1
    assert dev_per+val_per<1

    train_per = 1-dev_per-val_per

    description_column = []
    mpss_pl_column = []
    combined_list = []

    #Randomly shuffled the total file list
    for jira_workbook in work_sheet_list:
        jira_wb = openpyxl.load_workbook(jira_workbook)
        sheet_name = jira_wb.sheetnames[0]
        print "Sheet name is ", sheet_name
        work_sheet = jira_wb[sheet_name]
        description_column += work_sheet[DESCRIPTION_COL][1:-1]
        mpss_pl_column += work_sheet[MPSS_PL_COL][1:-1]

    combined_list = zip(mpss_pl_column, description_column)
    shuffled_list = random.sample(combined_list, len(description_column))
    #print shuffled_list
    total_cnt = len(shuffled_list)
    print "total cnt = %d"%(total_cnt)
    train_len = int(total_cnt * train_per)
    print "training samples' number is %d"%(train_len)
    dev_len = int(total_cnt * dev_per)
    print "dev samples' number is %d" % (dev_len)
    val_len = total_cnt - train_len - dev_len
    print "val samples' number is %d" % (val_len)

    train_list = shuffled_list[0:train_len]
    dev_list = shuffled_list[train_len:train_len+dev_len]
    val_list = shuffled_list[train_len+dev_len: total_cnt]

    return train_list, dev_list, val_list


def writeTfRecordData(cell_list, input_vocab, label_vocab, tf_record_file):
    writer = tf.python_io.TFRecordWriter(tf_record_file)
    for cell in cell_list:
        #fd = io.open(file, mode='r', encoding="ISO-8859-1")
        # When we store the data, first line is for labels
        #lines = fd.readlines()
        y_txt = cell[0].value
        #print y_txt
        x_txt = tokenize_helper(cell[1].value)
        #print x_lines
        if y_txt and x_txt:
            #x_txt = reduce(lambda x,y:x+y, x_lines)
            y = label_vocab.encode(y_txt) #No need to split y_txt since each mpss_pl is a whole phrase
            #print y
            x = input_vocab.encode_word_list(x_txt.split())
        else:
            #print "Either y_txt or x_lines is NULL"
            continue

        example = tf.train.Example(features=tf.train.Features(feature=
            {'y':tf.train.Feature(int64_list=tf.train.Int64List(value=[y])),
             'x':tf.train.Feature(int64_list=tf.train.Int64List(value=x))}))

        writer.write(example.SerializeToString())

    writer.close()


def tfrecord_main():

    #print jira_xlxs_list

    buildVocab(jira_xlxs_list)

    general_vocab = load_obj(FILEPATH, 'general_vocab')
    mpss_pl_vocab = load_obj(FILEPATH, 'mpss_pl_vocab')

    #general_vocab.word_distribution()

    general_vocab.filter_dictionary(lower_threshold=LOWER_DIC_FILTER_THRESHOLD)
    #mpss_pl_vocab.filter_dictionary(lower_threshold=LOWER_DIC_FILTER_THRESHOLD)

    general_vocab.reorder_dictionary()

    save_obj(FILEPATH, general_vocab, 'general_vocab')
    save_obj(FILEPATH, mpss_pl_vocab, 'mpss_pl_vocab')

    train_list, dev_list, val_list = split_train_dev_val(jira_xlxs_list, DEV_SAMPLE_PER, VAL_SAMPLE_PER)

    writeTfRecordData(train_list, general_vocab, mpss_pl_vocab, TRAIN_TFRECORD_FILE)
    writeTfRecordData(dev_list, general_vocab, mpss_pl_vocab, DEV_TFRECORD_FILE)
    writeTfRecordData(val_list, general_vocab, mpss_pl_vocab, VAL_TFRECORD_FILE)


if __name__ == '__main__':
    tfrecord_main()
